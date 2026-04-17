#!/usr/bin/env python3
"""
Re-import finished garment size charts from authoritative xlsx spec sheets.

Reads all xlsx files from the measurements folder, matches them to pattern slugs
in the database, and REPLACES the finished chart data with the xlsx values.

Usage:
  python3 scripts/fix-size-charts-from-xlsx.py                    # dry run
  python3 scripts/fix-size-charts-from-xlsx.py --apply             # write to DB
  python3 scripts/fix-size-charts-from-xlsx.py --apply --verbose   # write + details
"""

import argparse
import json
import os
import subprocess
import sys

import openpyxl

XLSX_DIR = "/Users/dr.contexter/Downloads/All Products Measurements/"

def get_secret(name):
    return subprocess.run(["pass", name], capture_output=True, text=True).stdout.strip()

DB_URL = get_secret("rosyspatterns/railway-database-url")

def psql_json(query):
    r = subprocess.run(["psql", DB_URL, "-t", "-A", "-c", f"SELECT json_agg(t) FROM ({query}) t;"], capture_output=True, text=True)
    raw = r.stdout.strip()
    return json.loads(raw) if raw and raw != "" else []

def psql_exec(query):
    r = subprocess.run(["psql", DB_URL, "-c", query], capture_output=True, text=True)
    if r.returncode != 0:
        print(f"  DB ERROR: {r.stderr.strip()}")
    return r.returncode == 0

SIZES = ['XXS', 'XS', 'S', 'M', 'L', 'XL', '2XL', '3XL', '4XL', '5XL']

# Map xlsx measurement labels to DB columns
LABEL_MAP = {
    'BUST': ('bust_cm', False),
    'BUST 1/2': ('bust_cm', True),
    'WAIST': ('waist_cm', False),
    'WAIST 1/2': ('waist_cm', True),
    'HIP': ('hip_cm', False),
    'HIP 1/2': ('hip_cm', True),
    'FULL LENGTH': ('full_length_cm', False),
    'SHOULDER': ('shoulder_cm', False),
    'SHOULDER 1/2': ('shoulder_cm', True),
    'SLEEVE LENGTH': ('sleeve_length_cm', False),
    'BOTTOM SWEEP': ('bottom_sweep_cm', False),
    'BOTTOM SWEEP 1/2': ('bottom_sweep_cm', True),
    'ZIPPER LENGTH': ('zipper_length_cm', False),
    'ZIPPER': ('zipper_length_cm', False),
}


def parse_xlsx(filepath):
    """Parse a size chart xlsx file. Returns {size: {column: value}}"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Find the header row with size names
    size_cols = {}
    for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=False), 1):
        vals = [str(c.value).strip().upper() if c.value else '' for c in row]
        found = 0
        for ci, v in enumerate(vals):
            if v in SIZES:
                found += 1
                size_cols[v] = ci
        if found >= 3:
            break

    if not size_cols:
        return None

    # Parse measurement rows
    data = {size: {} for size in size_cols}

    for row in ws.iter_rows(min_row=ri + 1, values_only=False):
        label = str(row[0].value).strip().upper() if row[0].value else ''
        if not label:
            continue

        mapping = LABEL_MAP.get(label)
        if not mapping:
            continue

        db_col, is_half = mapping

        for size, ci in size_cols.items():
            if ci < len(row) and row[ci].value is not None:
                try:
                    val = float(row[ci].value)
                    if is_half:
                        val *= 2
                    data[size][db_col] = val
                except (ValueError, TypeError):
                    pass

    wb.close()
    return data


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Actually write to DB")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    # Build slug lookup from catalog
    catalog = psql_json("SELECT pattern_slug, pattern_name FROM cs_pattern_catalog")
    slug_lookup = {}
    for c in catalog:
        name = c['pattern_name'].upper().strip()
        slug_lookup[name] = c['pattern_slug']

    total = 0
    fixed = 0
    skipped = 0
    errors = 0

    for fname in sorted(os.listdir(XLSX_DIR)):
        if not fname.endswith('.xlsx'):
            continue

        total += 1
        pattern_name = fname.replace(' SIZE CHART.xlsx', '').strip().upper()

        # Find slug
        slug = slug_lookup.get(pattern_name)
        if not slug:
            for cat_name, cat_slug in slug_lookup.items():
                if pattern_name in cat_name or cat_name in pattern_name:
                    slug = cat_slug
                    break

        if not slug:
            if args.verbose:
                print(f"  SKIP: {pattern_name} — no matching slug")
            skipped += 1
            continue

        # Parse xlsx
        try:
            data = parse_xlsx(os.path.join(XLSX_DIR, fname))
        except Exception as e:
            print(f"  ERROR: {fname} — {e}")
            errors += 1
            continue

        if not data:
            if args.verbose:
                print(f"  SKIP: {fname} — can't parse sizes")
            skipped += 1
            continue

        if args.verbose:
            print(f"\n  {slug} ({pattern_name}):")

        cols = ['bust_cm', 'waist_cm', 'hip_cm', 'full_length_cm',
                'shoulder_cm', 'sleeve_length_cm', 'bottom_sweep_cm', 'zipper_length_cm']

        for size, measurements in sorted(data.items(), key=lambda x: SIZES.index(x[0]) if x[0] in SIZES else 99):
            if not measurements:
                continue

            if args.verbose:
                vals = ', '.join(f"{k}={v}" for k, v in sorted(measurements.items()) if v is not None)
                print(f"    {size}: {vals}")

            if args.apply:
                # Build UPSERT
                set_parts = []
                val_parts = []
                for col in cols:
                    val = measurements.get(col)
                    if val is not None:
                        set_parts.append(f"{col} = {val}")
                        val_parts.append(str(val))
                    else:
                        val_parts.append("NULL")

                insert_vals = f"'{slug}', 'finished', '{size}', " + ", ".join(val_parts)
                update_set = ", ".join(set_parts) if set_parts else None

                if update_set:
                    sql = f"""
                        INSERT INTO cs_pattern_size_charts
                            (pattern_slug, chart_type, size, {', '.join(cols)})
                        VALUES ({insert_vals})
                        ON CONFLICT (pattern_slug, chart_type, size)
                        DO UPDATE SET {update_set};
                    """
                    psql_exec(sql)

        fixed += 1

    print(f"\n{'='*60}")
    print(f"RESULTS: {total} xlsx files")
    print(f"  Fixed/updated: {fixed}")
    print(f"  Skipped:       {skipped}")
    print(f"  Errors:        {errors}")
    if not args.apply:
        print(f"\n  DRY RUN — use --apply to write to DB")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
